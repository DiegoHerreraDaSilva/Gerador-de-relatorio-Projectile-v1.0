"""Automação do preenchimento de Horas Faturadas e KPI (Dias) no Painel de
Gerência a partir dos e-mails do gerente de engenharia (Alberto) pros
clientes, com a caixa `agente.reunioes@schwaben.com.br` em cópia.

Fluxo (`process_new_emails`, chamado periodicamente pelo loop de polling em
`main.py`): busca mensagens novas na caixa via Microsoft Graph (client
credentials, sem interação de usuário) → baixa o(s) anexo(s) `.xlsx`/`.pdf`
→ lê o projeto e resolve o valor de "Total de horas Mês/Ano:" (ver
`resolve_total_hours`/`read_pdf_report_data` abaixo) → faz fuzzy match do
nome do projeto contra o Projectile → calcula dias úteis entre o fechamento
do mês do relatório e o envio do e-mail → grava uma amostra em
`backend/data/management_kpi.json` via `management.append_project_kpi_sample`.

Não usa `data_only=True`/valor em cache do Excel: os relatórios `.xlsx` deste
app são gerados via ZIP/XML manual (`generator.py`, nunca
`openpyxl.save()`), então a fórmula nunca tem um `<v>` calculado embutido —
mesmo que o anexo já tenha sido aberto no Excel por alguém no meio do
caminho, não há garantia disso. `resolve_total_hours` por isso segue a
cadeia de fórmulas manualmente, reconhecendo só as 3 formas exatas que
`generator._build_groups_xml` escreve.

O `.pdf` não tem fórmula nem célula pra ler — `pdf_generator.py` grava os
mesmos dados (código/nome do projeto, mês, total já calculado, pacote de
trabalho coberto) como metadado do próprio arquivo (Info dictionary,
`PDF_METADATA_KEY`), lido de volta aqui por `read_pdf_report_data`. Quando um
mesmo relatório chega nos dois formatos na mesma mensagem (transição pro
novo formato, ou envio duplicado por hábito), só UM é processado — ver
`fetch_report_attachments` — pra não somar a mesma hora faturada duas vezes.
"""
from __future__ import annotations

import base64
import binascii
import difflib
import json
import os
import re
import tempfile
import zipfile
from datetime import date, datetime, timedelta, timezone

import msal
import requests
from openpyxl import load_workbook
from pypdf import PdfReader

from . import management
from .generator import parse_month_label, count_business_days, HIDDEN_HELPER_COL
from .pdf_generator import PDF_METADATA_KEY
from .projectile_db import ProjectileDbError, fetch_all_projects

GRAPH_BASE = "https://graph.microsoft.com/v1.0"
_MAX_FORMULA_DEPTH = 10
_CELL_REF_RE = re.compile(r"^([A-Z]+)(\d+)$")


class EmailIngestError(RuntimeError):
    """Falha ao consultar o Graph, ler um anexo ou resolver a fórmula de
    "Total de horas" — nunca um erro do usuário, sempre um problema de
    configuração/anexo/template que precisa de atenção manual."""


def _require_env(name: str) -> str:
    value = os.environ.get(name)
    if not value:
        raise EmailIngestError(
            f"Configuração ausente para a automação de e-mail: defina {name} no .env "
            "(veja .env.example)."
        )
    return value


# App confidencial único em nível de módulo — o MSAL já cacheia o token de
# acesso internamente entre chamadas, então não é preciso gerenciar expiração
# aqui manualmente (mesmo espírito de reaproveitamento de
# `projectile_db._get_connection`, adaptado pra um client HTTP em vez de uma
# conexão de banco).
_msal_app: msal.ConfidentialClientApplication | None = None


def _get_msal_app() -> msal.ConfidentialClientApplication:
    global _msal_app
    if _msal_app is None:
        tenant_id = _require_env("AZURE_TENANT_ID")
        client_id = _require_env("AZURE_CLIENT_ID")
        client_secret = _require_env("AZURE_CLIENT_SECRET")
        _msal_app = msal.ConfidentialClientApplication(
            client_id,
            authority=f"https://login.microsoftonline.com/{tenant_id}",
            client_credential=client_secret,
        )
    return _msal_app


def _get_graph_token() -> str:
    app = _get_msal_app()
    result = app.acquire_token_silent(["https://graph.microsoft.com/.default"], account=None)
    if not result:
        result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])
    if "access_token" not in result:
        raise EmailIngestError(
            f"Falha ao autenticar no Microsoft Graph: {result.get('error_description') or result}"
        )
    return result["access_token"]


def _graph_get(url: str, params: dict | None = None) -> dict:
    """Isola toda chamada HTTP ao Graph num único ponto — facilita mockar a
    orquestração (`process_new_emails`) em teste sem precisar de credencial
    real, e centraliza o tratamento de erro."""
    token = _get_graph_token()
    try:
        resp = requests.get(
            url, params=params, headers={"Authorization": f"Bearer {token}"}, timeout=30
        )
        resp.raise_for_status()
        return resp.json()
    except (requests.RequestException, ValueError) as e:
        # ValueError cobre json.JSONDecodeError: um 2xx com corpo não-JSON (ex:
        # página de erro HTML de um proxy corporativo) não é RequestException,
        # mas ainda é uma falha de infra, não um bug — mesmo tratamento genérico.
        raise EmailIngestError(f"Falha ao consultar o Microsoft Graph: {e}") from e


def _graph_post(url: str, json_body: dict) -> None:
    """Mesmo isolamento de `_graph_get`, mas pra chamadas POST sem corpo de
    resposta relevante (ex: `sendMail`, que devolve 202 vazio em sucesso)."""
    token = _get_graph_token()
    try:
        resp = requests.post(
            url, json=json_body, headers={"Authorization": f"Bearer {token}"}, timeout=30
        )
        resp.raise_for_status()
    except requests.RequestException as e:
        # o corpo da resposta de erro do Graph costuma ter o motivo real
        # (ex: "não encontrei a caixa X", "sem permissão pra enviar como Y")
        # — inclui aqui pra ajudar a diagnosticar, já que quem chama isso é
        # sempre uma ação explícita do usuário (enviar relatório), não uma
        # automação silenciosa de fundo.
        detail = ""
        if e.response is not None:
            try:
                detail = f" — {e.response.json().get('error', {}).get('message', '')}"
            except ValueError:
                detail = f" — {e.response.text[:200]}"
        raise EmailIngestError(f"Falha ao enviar e-mail pelo Microsoft Graph{detail}") from e


def fetch_new_messages() -> list[dict]:
    """Mensagens do Alberto na caixa `GRAPH_MAILBOX` com anexo, ainda não
    processadas (ver `management.is_message_processed`). Filtra no próprio
    Graph (`$filter`) pra não baixar corpo/metadado de mensagens irrelevantes.

    Sem `$orderby`: combinado com este `$filter` (que usa uma propriedade
    aninhada, `from/emailAddress/address`), o Graph devolve 400
    "InefficientFilter" ("restriction or sort order is too complex") —
    confirmado testando ao vivo contra a caixa real. Não precisamos de ordem
    nenhuma aqui (a idempotência é por `message_id`, não por posição), então
    a solução é simplesmente não pedir ordenação, em vez de contornar com o
    header `ConsistencyLevel: eventual` (que muda semântica de contagem e
    paginação sem necessidade real neste caso)."""
    mailbox = _require_env("GRAPH_MAILBOX")
    sender = _require_env("ALBERTO_EMAIL")
    url = f"{GRAPH_BASE}/users/{mailbox}/messages"
    params = {
        "$filter": f"from/emailAddress/address eq '{sender}' and hasAttachments eq true",
        "$select": "id,subject,receivedDateTime,from,internetMessageHeaders",
        "$top": "50",
    }
    messages: list[dict] = []
    while url:
        data = _graph_get(url, params)
        for msg in data.get("value", []):
            if not management.is_message_processed(msg["id"]):
                messages.append(msg)
        url = data.get("@odata.nextLink")
        params = None  # já embutido no nextLink
    return messages


_MAX_ATTACHMENT_BYTES = 25 * 1024 * 1024  # 25 MB
_MAX_UNCOMPRESSED_ATTACHMENT_BYTES = 200 * 1024 * 1024  # 200 MB


_REPORT_EXTENSIONS = (".xlsx", ".pdf")
# ordem de preferência quando o mesmo relatório chega nos dois formatos na
# mesma mensagem (ver dedup por nome-base logo abaixo) — .xlsx é o caminho
# mais testado, fica como formato "principal" nesse empate.
_EXT_PRIORITY = {".xlsx": 0, ".pdf": 1}


def fetch_report_attachments(message_id: str) -> tuple[list[str], list[str]]:
    """Baixa os anexos `.xlsx`/`.pdf` da mensagem pra arquivos temporários
    (mesmo padrão de `parser.py`/`generator.py`, que também usam `tempfile`)
    e devolve `(paths, skip_reasons)` — caminhos utilizáveis e,
    separadamente, o motivo de cada anexo que foi pulado.

    Duas camadas de defesa contra "zip bomb" pro `.xlsx`, mesma spec de
    `main._reject_if_oversized_xlsx` (duplicado aqui de propósito: ponto de
    entrada diferente, bytes já vêm decodificados de base64 — ver
    CLAUDE.md): (1) tamanho do payload decodificado (também aplicado ao
    `.pdf` — não tem estrutura de zip pra "bomba", mas ainda precisa de um
    teto de tamanho) e (2) tamanho DESCOMPRIMIDO do conteúdo do zip, só pro
    `.xlsx`. Arquivo `.xlsx` que não abre como zip não é rejeitado aqui,
    mesma escolha de `main.py` — deixa a validação existente mais adiante
    (`resolve_total_hours`/`read_project_identity`, via openpyxl) reportar
    como anexo inválido/corrompido.

    Cada anexo é isolado no próprio try/except dentro do loop: um anexo
    grande demais, com Base64 malformado ou que falhe de qualquer outra
    forma nunca aborta a mensagem inteira — só é pulado (motivo registrado
    em `skip_reasons`, ver `process_new_emails`) — e nunca deixa órfão em
    disco o temp file de um anexo anterior já baixado com sucesso, já que a
    limpeza de um anexo que falhou acontece aqui mesmo, antes de seguir pro
    próximo.

    Se o mesmo relatório chegar em `.xlsx` E `.pdf` na mesma mensagem (nome
    de arquivo igual, só a extensão muda — garantido por
    `main._sanitized_file_name`, o mesmo `file_name` gera os dois formatos),
    só UM dos dois é devolvido: `billed_hours` é SOMADO entre amostras
    (`management.compute_monthly_kpis`), então processar os dois criaria 2
    amostras pro mesmo relatório e dobraria a hora faturada. Não é tratado
    como erro (não entra em `skip_reasons`) — é o comportamento esperado
    durante uma transição de formato."""
    mailbox = _require_env("GRAPH_MAILBOX")
    data = _graph_get(f"{GRAPH_BASE}/users/{mailbox}/messages/{message_id}/attachments")
    downloaded: list[tuple[str, str]] = []  # (tmp_path, nome original) — só pro dedup por nome-base abaixo
    skip_reasons: list[str] = []
    for att in data.get("value", []):
        name = att.get("name") or ""
        content_bytes = att.get("contentBytes")
        ext = os.path.splitext(name)[1].lower()
        if not content_bytes or ext not in _REPORT_EXTENSIONS:
            continue

        try:
            decoded = base64.b64decode(content_bytes)
        except (binascii.Error, ValueError) as e:
            skip_reasons.append(f"anexo '{name}' com conteúdo Base64 inválido: {e}")
            continue

        if len(decoded) > _MAX_ATTACHMENT_BYTES:
            skip_reasons.append(f"anexo '{name}' excede o limite de tamanho (25 MB)")
            continue

        tmp_path: str | None = None
        try:
            with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as tmp:
                tmp.write(decoded)
                tmp_path = tmp.name

            if ext == ".xlsx":
                try:
                    with zipfile.ZipFile(tmp_path) as zf:
                        total_uncompressed = sum(info.file_size for info in zf.infolist())
                    if total_uncompressed > _MAX_UNCOMPRESSED_ATTACHMENT_BYTES:
                        skip_reasons.append(
                            f"anexo '{name}' excede o limite de tamanho descomprimido (200 MB)"
                        )
                        os.remove(tmp_path)
                        continue
                except zipfile.BadZipFile:
                    pass  # não é um zip válido — deixa a validação existente adiante rejeitar

            downloaded.append((tmp_path, name))
        except Exception as e:
            if tmp_path is not None:
                try:
                    os.remove(tmp_path)
                except OSError:
                    pass
            skip_reasons.append(f"falha ao baixar anexo '{name}': {e}")

    return _dedupe_by_stem(downloaded), skip_reasons


def _dedupe_by_stem(downloaded: list[tuple[str, str]]) -> list[str]:
    """Isolada de `fetch_report_attachments` pra ser testável sem precisar
    mockar o Graph: recebe `(tmp_path, nome_original)` de cada anexo já
    baixado e devolve só os `tmp_path` que sobrevivem ao dedup por nome-base
    (ver docstring de `fetch_report_attachments`) — o(s) `tmp_path` perdedor
    já é removido do disco aqui, quem chama não precisa limpar de novo."""
    by_stem: dict[str, tuple[str, str]] = {}
    for path, name in downloaded:
        stem = os.path.splitext(name)[0].strip().casefold()
        ext = os.path.splitext(name)[1].lower()
        existing = by_stem.get(stem)
        if existing is None:
            by_stem[stem] = (path, name)
            continue
        existing_path, existing_name = existing
        existing_ext = os.path.splitext(existing_name)[1].lower()
        if _EXT_PRIORITY.get(ext, 99) < _EXT_PRIORITY.get(existing_ext, 99):
            try:
                os.remove(existing_path)
            except OSError:
                pass
            by_stem[stem] = (path, name)
        else:
            try:
                os.remove(path)
            except OSError:
                pass

    return [path for path, _ in by_stem.values()]


def _resolve_cell(ws, ref: str, seen: set[str], depth: int = 0) -> float:
    """Segue a cadeia de fórmulas escrita por `generator._build_groups_xml`
    até um número. Reconhece só as 3 formas que esse template gera:
    `=SUM(ref,ref,...)`, referência simples a outra célula (`=E12`), e
    produto de duas células (`=E12*F12`). Qualquer outra forma é um sinal de
    que o template mudou — falha alto em vez de devolver um número errado
    silenciosamente."""
    if depth > _MAX_FORMULA_DEPTH or ref in seen:
        raise EmailIngestError(f"Fórmula circular ou profunda demais ao resolver {ref}.")
    seen = seen | {ref}

    value = ws[ref].value
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if not isinstance(value, str) or not value.startswith("="):
        raise EmailIngestError(f"Valor inesperado (não numérico, não fórmula) em {ref}: {value!r}")

    formula = value[1:]

    sum_match = re.fullmatch(r"SUM\(([^()]+)\)", formula)
    if sum_match:
        refs = [r.strip() for r in sum_match.group(1).split(",") if r.strip()]
        return sum(_resolve_cell(ws, r, seen, depth + 1) for r in refs)

    product_match = re.fullmatch(r"([A-Z]+\d+)\*([A-Z]+\d+)", formula)
    if product_match:
        return _resolve_cell(ws, product_match.group(1), seen, depth + 1) * _resolve_cell(
            ws, product_match.group(2), seen, depth + 1
        )

    if _CELL_REF_RE.fullmatch(formula):
        return _resolve_cell(ws, formula, seen, depth + 1)

    raise EmailIngestError(f"Forma de fórmula não reconhecida em {ref}: ={formula}")


def _find_total_row(ws) -> tuple[int, str]:
    """Acha a linha "Total de horas <mês/ano>:" (coluna B) — devolve (número
    da linha, month_label). Compartilhado por `resolve_total_hours` e
    `read_pacote_scope` (mesma linha-âncora — a linha logo abaixo é
    `bruto_row`, ver `generator._build_totals_row`)."""
    label_re = re.compile(r"^Total de horas\s+(.+?):$")
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            match = label_re.match(cell.value.strip())
            if match:
                return cell.row, match.group(1).strip()
    raise EmailIngestError("Não encontrei a célula 'Total de horas ...:' no anexo.")


def resolve_total_hours(xlsx_path: str) -> tuple[float, str]:
    """Acha a linha "Total de horas <mês/ano>:" na coluna B e resolve a
    fórmula da coluna C na mesma linha. Devolve (horas, month_label) — o
    `month_label` (ex: "Julho/2026") vem do próprio texto da label, é a
    competência do relatório, não a data de hoje nem a data de envio."""
    wb = load_workbook(xlsx_path, data_only=False)
    ws = wb.active
    total_row, month_label = _find_total_row(ws)
    hours = _resolve_cell(ws, f"C{total_row}", set())
    return round(hours, 3), month_label


def read_pacote_scope(xlsx_path: str) -> str | None:
    """Lê a marca oculta gravada por `generator._build_totals_row`
    (HIDDEN_HELPER_COL na linha logo abaixo de "Total de horas ...:", ==
    `bruto_row`) — célula vazia/ausente significa que o relatório cobre o
    projeto inteiro; um texto significa que cobre só aquele pacote de
    trabalho. Usada por `management.compute_monthly_kpis` pra decidir status
    "enviado"/"parcial" por projeto em vez de marcar o projeto inteiro como
    enviado a partir de um relatório de 1 pacote só."""
    wb = load_workbook(xlsx_path, data_only=False)
    ws = wb.active
    total_row, _ = _find_total_row(ws)
    value = str(ws[f"{HIDDEN_HELPER_COL}{total_row + 1}"].value or "").strip()
    return value or None


def read_project_identity(xlsx_path: str) -> tuple[str, str]:
    """Lê `project_code` (B8) e `project_name` (C9), gravados por
    `generator.generate_report` em todo relatório emitido pelo app."""
    wb = load_workbook(xlsx_path, data_only=False)
    ws = wb.active
    project_code = str(ws["B8"].value or "").strip()
    project_name = str(ws["C9"].value or "").strip()
    if not project_name:
        raise EmailIngestError("Não encontrei o nome do projeto (célula C9) no anexo.")
    return project_code, project_name


def read_pdf_report_data(pdf_path: str) -> dict:
    """Equivalente, pro `.pdf`, de `resolve_total_hours` +
    `read_project_identity` + `read_pacote_scope` juntos: o `.pdf` não tem
    fórmula nem célula pra ler, então `pdf_generator._embed_report_metadata`
    já grava tudo isso de uma vez como metadado do arquivo (Info dictionary,
    `PDF_METADATA_KEY`) no momento da geração — aqui só lê de volta e valida
    que veio completo. Um PDF que não foi gerado por este app (ou teve o
    metadado apagado/editado) não tem essa chave ou vem incompleta — mesmo
    tratamento de erro que um `.xlsx` sem a célula/label esperada."""
    reader = PdfReader(pdf_path)
    raw = (reader.metadata or {}).get(PDF_METADATA_KEY)
    if not raw:
        raise EmailIngestError(
            "Não encontrei os metadados do relatório no PDF anexado — ele não foi "
            "gerado por este app, ou o metadado foi removido."
        )
    try:
        data = json.loads(raw)
    except (TypeError, ValueError) as e:
        raise EmailIngestError(f"Metadados do PDF corrompidos: {e}") from e
    if not data.get("project_name") or not data.get("month_label") or data.get("total_hours") is None:
        raise EmailIngestError("Metadados do PDF incompletos (relatório gerado por uma versão antiga do app?).")
    return data


def match_project(report_project_name: str, candidates: dict[str, str]) -> tuple[str, str, float]:
    """Fuzzy match do texto livre do relatório contra `tproject.pDescription`
    — decisão do usuário: sempre aplica o melhor candidato disponível, mesmo
    com pouca confiança (sem fila de revisão manual). `match_score` fica
    salvo na amostra pra auditoria via o endpoint de diagnóstico."""
    if not candidates:
        raise EmailIngestError("Nenhum projeto cadastrado no Projectile para comparar.")
    names = list(candidates.values())
    ids_by_name: dict[str, str] = {}
    for pid, name in candidates.items():
        ids_by_name.setdefault(name, pid)

    best_name = None
    best_score = 0.0
    for name in names:
        score = difflib.SequenceMatcher(None, report_project_name.casefold(), name.casefold()).ratio()
        if score > best_score:
            best_score, best_name = score, name
    return ids_by_name[best_name], best_name, round(best_score, 3)


def _month_end(year: int, month: int) -> date:
    next_month = date(year + 1, 1, 1) if month == 12 else date(year, month + 1, 1)
    return next_month - timedelta(days=1)


def compute_business_days_elapsed(month_label: str, sent_at: datetime) -> int:
    """Dias úteis entre o último dia do mês do relatório (competência lida da
    própria label "Total de horas <mês/ano>:") e a data de envio do e-mail."""
    parsed = parse_month_label(month_label)
    if parsed is None:
        raise EmailIngestError(f"Não reconheço o formato de competência: {month_label!r}")
    year, month = parsed
    anchor = _month_end(year, month)
    return count_business_days(anchor, sent_at.date())


def _sender_failed_dmarc(msg: dict) -> bool:
    """Defesa em profundidade além do `$filter` por `from/emailAddress/address`
    em `fetch_new_messages` — esse filtro só reflete o header `From:` como
    entregue, que pode ser forjado se a política DMARC do tenant não
    rejeitar/colocar em quarentena e-mail forjado do domínio. Procura o
    header `Authentication-Results` (exposto pelo Graph via
    `internetMessageHeaders`, incluído no `$select` de `fetch_new_messages`).

    Falha fechado (`True`) só numa falha DMARC CONFIRMADA (`dmarc=fail`
    explícito no valor do header). Falha aberto (`False`) em qualquer
    ambiguidade — header ausente, ou presente sem esse valor exato
    (`dmarc=pass`, `dmarc=bestguesspass`, `dmarc=none`) — porque e-mail
    intra-organização comumente não carrega esse header dependendo da
    config do tenant/Exchange, e esta automação já foi validada
    funcionando ponta a ponta e alimenta KPIs de negócio reais: não deve
    parar de processar tudo silenciosamente por causa de um header que nem
    sempre está presente nesse cenário comum."""
    for header in msg.get("internetMessageHeaders") or []:
        name = (header.get("name") or "").strip().lower()
        if name == "authentication-results":
            if "dmarc=fail" in (header.get("value") or "").lower():
                return True
    return False


_ATTACHMENT_CONTENT_TYPES = {
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}


def send_report_email(
    sender_email: str,
    to_email: str,
    subject: str,
    body_text: str,
    attachments: list[tuple[str, bytes]],
) -> None:
    """Envia relatório(s) por e-mail via Microsoft Graph, "como" o usuário
    logado (`sender_email` — vem da sessão, `auser.rEmail`, nunca digitado
    pelo cliente, pra não dar pra forjar remetente) — `POST
    /users/{sender_email}/sendMail`. `attachments` é uma lista de
    `(nome_do_arquivo, bytes)` — cada um vira um anexo (`.xlsx` ou `.pdf`,
    ver `_ATTACHMENT_CONTENT_TYPES`) SEPARADO no
    mesmo e-mail (nunca zipado junto — decisão do usuário: mais de um
    relatório pro mesmo destinatário pode ir num único e-mail, cada `.xlsx`
    continua individual). Sempre copia `GRAPH_MAILBOX` (a caixa do agente),
    fixo e não editável pelo usuário, pra automação de KPI
    (`process_new_emails`) capturar o envio do mesmo jeito que já captura
    os e-mails do Alberto — inclusive quando o e-mail tem vários anexos
    (já suportado desde a correção de mensagens com múltiplos relatórios).

    Exige a permissão de aplicação `Mail.Send` no Graph (além do `Mail.Read`
    já usado pra ler a caixa do agente) — sem uma Application Access Policy
    restringindo o escopo, essa permissão deixa o app capaz de mandar e-mail
    "como" qualquer caixa do tenant, não só a de quem está logado; ver
    CLAUDE.md/README pra essa configuração externa (ação de admin do Azure,
    fora do código)."""
    mailbox = _require_env("GRAPH_MAILBOX")
    url = f"{GRAPH_BASE}/users/{sender_email}/sendMail"
    body = {
        "message": {
            "subject": subject,
            "body": {"contentType": "Text", "content": body_text},
            "toRecipients": [{"emailAddress": {"address": to_email}}],
            "ccRecipients": [{"emailAddress": {"address": mailbox}}],
            "attachments": [
                {
                    "@odata.type": "#microsoft.graph.fileAttachment",
                    "name": name,
                    "contentType": _ATTACHMENT_CONTENT_TYPES.get(
                        name.rsplit(".", 1)[-1].lower(), "application/octet-stream"
                    ),
                    "contentBytes": base64.b64encode(content).decode("ascii"),
                }
                for name, content in attachments
            ],
        },
        "saveToSentItems": True,
    }
    _graph_post(url, body)


def process_new_emails() -> dict:
    """Orquestra um ciclo de polling. Chamado periodicamente pelo loop em
    `main.py` (`_poll_emails_loop`) e também sob demanda pelo botão
    "Verificar horas faturadas" do painel (`POST /management/kpis/check-emails`).
    Idempotente por `message_id` — seguro de chamar repetidamente, inclusive
    após um restart do `--reload` ou disparado manualmente enquanto o
    polling automático também está rodando. Devolve um resumo (contagens)
    pra dar feedback imediato de quem chamou sob demanda."""
    messages = fetch_new_messages()
    summary = {"messages_found": len(messages), "samples_added": 0, "duplicates_found": 0, "skipped": 0}
    if not messages:
        return summary

    candidates = fetch_all_projects()

    for msg in messages:
        message_id = msg["id"]
        received_at = msg.get("receivedDateTime") or ""
        try:
            sent_at = datetime.fromisoformat(received_at.replace("Z", "+00:00"))
        except ValueError:
            sent_at = datetime.now(timezone.utc)

        if _sender_failed_dmarc(msg):
            management.append_skipped_message(
                message_id, received_at, "Falha na verificação DMARC do remetente — possível spoofing"
            )
            summary["skipped"] += 1
            continue

        try:
            attachments, download_skip_reasons = fetch_report_attachments(message_id)
            if not attachments and not download_skip_reasons:
                management.append_skipped_message(message_id, received_at, "anexo .xlsx/.pdf não encontrado")
                summary["skipped"] += 1
                continue

            # Alberto pode mandar mais de um relatório (projetos diferentes)
            # no mesmo e-mail — processa TODOS os anexos, não só o primeiro,
            # cada um virando sua própria amostra (mesmo `email_message_id`,
            # projetos/horas/dias possivelmente diferentes entre si).
            sender = (msg.get("from") or {}).get("emailAddress", {}).get("address", "")
            attachment_errors: list[str] = list(download_skip_reasons)
            try:
                for report_path in attachments:
                    try:
                        if report_path.lower().endswith(".pdf"):
                            pdf_data = read_pdf_report_data(report_path)
                            billed_hours = round(float(pdf_data["total_hours"]), 3)
                            month_label = pdf_data["month_label"]
                            report_project_name = pdf_data["project_name"]
                            pacote_scope = pdf_data.get("pacote_scope")
                        else:
                            billed_hours, month_label = resolve_total_hours(report_path)
                            _, report_project_name = read_project_identity(report_path)
                            pacote_scope = read_pacote_scope(report_path)
                        project_id, project_name, score = match_project(report_project_name, candidates)
                        business_days = compute_business_days_elapsed(month_label, sent_at)
                        parsed_month = parse_month_label(month_label)
                        month_key = f"{parsed_month[0]:04d}-{parsed_month[1]:02d}"

                        is_duplicate = management.append_project_kpi_sample({
                            "email_message_id": message_id,
                            "received_at": received_at,
                            "sender": sender,
                            "report_project_text": report_project_name,
                            "project_id": project_id,
                            "project_name": project_name,
                            "match_score": score,
                            "month": month_key,
                            "billed_hours": billed_hours,
                            "business_days": business_days,
                            "pacote_scope": pacote_scope,
                        })
                        if is_duplicate:
                            summary["duplicates_found"] += 1
                        else:
                            summary["samples_added"] += 1
                    except (EmailIngestError, ProjectileDbError) as e:
                        attachment_errors.append(str(e))
                    except Exception as e:
                        # Anexo corrompido/não é realmente um .xlsx ou .pdf válido
                        # (zipfile.BadZipFile, openpyxl InvalidFileException,
                        # pypdf.errors.PdfReadError etc.) não é
                        # EmailIngestError/ProjectileDbError — sem este except,
                        # um único anexo malformado derrubava o processamento de TODA
                        # a mensagem (ver except mais abaixo) em vez de só marcar esse
                        # anexo como pulado e seguir com os demais.
                        attachment_errors.append(f"Anexo inválido ou corrompido: {e}")
            finally:
                for path in attachments:
                    try:
                        os.remove(path)
                    except OSError:
                        pass

            if attachment_errors:
                management.append_skipped_message(message_id, received_at, "; ".join(attachment_errors))
                summary["skipped"] += 1
        except (EmailIngestError, ProjectileDbError) as e:
            management.append_skipped_message(message_id, received_at, str(e))
            summary["skipped"] += 1
        except Exception as e:
            # Falha inesperada só nesta mensagem (ex: Base64 corrompido em
            # fetch_report_attachments, erro de I/O ao gravar management_kpi.json) não
            # pode derrubar o ciclo inteiro: sem isto, as mensagens seguintes do
            # mesmo polling nunca seriam processadas, e como esta mensagem nunca é
            # marcada como processada, ela travaria a automação tentando de novo a
            # cada ciclo, para sempre.
            management.append_skipped_message(message_id, received_at, f"Erro inesperado: {e}")
            summary["skipped"] += 1

    return summary
