"""Leitura e agrupamento do relatório exportado do Projectile."""
from __future__ import annotations

import re
from dataclasses import dataclass, field

import openpyxl


@dataclass
class Activity:
    description: str
    hours: float


@dataclass
class Group:
    name: str
    activities: list[Activity] = field(default_factory=list)

    @property
    def total_hours(self) -> float:
        return round(sum(a.hours for a in self.activities), 3)


@dataclass
class WorkPackage:
    key: str
    project_name: str
    groups: list[Group] = field(default_factory=list)


@dataclass
class RowIssue:
    row: int
    reason: str
    message: str


# variantes de travessão/hífen que aparecem em exports reais no lugar do hífen ASCII
# comum ("-") — normalizadas antes de separar os segmentos de "Pacote de Trabalho".
_DASH_VARIANTS = ["–", "—", "−", "－"]  # en dash, em dash, minus sign, fullwidth hyphen


def _extract_package_key(raw_value) -> str | None:
    """Extrai o identificador do projeto a partir da coluna "Pacote de Trabalho".

    O valor vem no formato "<código>-<seq> Legislation Package - <Projeto> - <resto>"
    (ex: "1546.6.4-002 Legislation Package - Sangam - Cabina Bruta - ..."). O separador
    dos segmentos é " - " (espaço-hífen-espaço), não hífen puro — usar hífen puro
    quebraria nomes de projeto que já contêm hífen, como "Para-barro". O nome do
    projeto é o 2º segmento — é isso que agrupa várias linhas de "Pacote de Trabalho"
    diferentes (6.4-001, 6.4-002, ...) num único relatório ("Sangam").

    Retorna None quando o valor NÃO segue essa estrutura (menos de 2 segmentos depois
    de separar por " - ") — ex: "N/A", "TBD", "-", ou qualquer texto livre. Antes essa
    situação caía num fallback ingênuo que devolvia o texto bruto como se fosse um nome
    de projeto confiável; o problema é que duas linhas de projetos genuinamente
    diferentes podem compartilhar o mesmo texto de preenchimento genérico (ambas com
    "N/A", por exemplo) e acabariam sendo mescladas num único pacote sem nenhum aviso.
    Devolver None força o chamador a tratar como "não identificado" (chave própria por
    linha + aviso) em vez de confiar cegamente no texto.
    """
    if not raw_value:
        return None
    # separador tolera espaçamento irregular (às vezes vem "- " ou " -" só de um lado,
    # ou espaço duplo), normaliza non-breaking space (\xa0) e variantes de travessão
    # unicode (en/em dash, sinal de menos, hífen fullwidth), comuns em exports do Excel.
    normalized = str(raw_value).replace("\xa0", " ")
    for dash in _DASH_VARIANTS:
        normalized = normalized.replace(dash, "-")
    # o padrão exige espaço em pelo menos um dos lados do hífen (não precisa ser dos
    # dois) — só assim ele conta como separador de segmento; um hífen sem espaço
    # nenhum ao redor (ex: "Para-barro") nunca é separador. Um padrão assimétrico
    # (só exigindo espaço de um lado fixo) deixaria passar despercebido um separador
    # real que só tem espaço do outro lado, quebrando o texto no lugar errado.
    parts = re.split(r"\s+-\s*|\s*-\s+", normalized)
    if len(parts) >= 2:
        return parts[1].strip() or None
    return None


def _parse_hs_value(value) -> float:
    """Converte o valor de Hs para float, aceitando tanto ponto quanto vírgula como
    separador decimal — exports variam conforme a configuração regional de quem gerou
    a planilha, e um valor como "2,5" é um apontamento real, não deveria ser tratado
    como texto de rodapé só porque float() sozinho não entende vírgula.
    """
    try:
        return float(value)
    except (TypeError, ValueError):
        pass
    if isinstance(value, str):
        try:
            return float(value.strip().replace(",", "."))
        except ValueError:
            pass
    raise ValueError(f"valor de Hs não numérico: {value!r}")


# Chave interna de pacote usada no modo "relatório único" (tudo cai num só
# WorkPackage) antes de o nome real ficar conhecido — nunca deve vazar pra
# tela/relatório final (ver o rename feito via _PackageAccumulator.rename_package
# ao final de parse_projectile_export/group_hours). Compartilhada entre os dois
# formatos (Excel e MySQL, ver backend/app/projectile_db.py) por serem, nesse
# ponto, a mesma regra: "sem split por pacote, um WorkPackage só".
SINGLE_PACKAGE_KEY = "__single__"


class _PackageAccumulator:
    """Acumula apontamentos (de qualquer origem — linha de Excel ou linha de
    banco) em WorkPackage/Group/Activity: agrupa por pacote e, dentro de cada
    pacote, por prefixo (Group), somando horas de atividades com a mesma
    descrição (case-insensitive) em vez de duplicá-las.

    Extraído porque `parser.parse_projectile_export` e
    `projectile_db.group_hours` implementavam essa acumulação de forma
    idêntica — só a resolução de package_key/package_name/prefixo/descrição/
    horas antes de chamar `add_activity` difere entre os dois (validação de
    linha incompleta e conversão de Hs texto-livre só existem no Excel;
    dedupe de HTML entities só existe no MySQL).
    """

    def __init__(self) -> None:
        self._packages: dict[str, WorkPackage] = {}
        self._package_order: list[str] = []
        self._groups_by_package: dict[str, dict[str, Group]] = {}
        self._group_order_by_package: dict[str, list[str]] = {}

    def add_activity(
        self, package_key: str, package_name: str, group_name: str, description: str, hours: float
    ) -> None:
        if package_key not in self._packages:
            self._packages[package_key] = WorkPackage(key=package_key, project_name=package_name)
            self._package_order.append(package_key)
            self._groups_by_package[package_key] = {}
            self._group_order_by_package[package_key] = []

        groups = self._groups_by_package[package_key]
        group_order = self._group_order_by_package[package_key]
        if group_name not in groups:
            groups[group_name] = Group(name=group_name)
            group_order.append(group_name)
        group = groups[group_name]

        existing = next(
            (a for a in group.activities if a.description.casefold() == description.casefold()), None
        )
        if existing:
            existing.hours = round(existing.hours + hours, 3)
        else:
            group.activities.append(Activity(description=description, hours=hours))

    def rename_package(self, package_key: str, name: str) -> None:
        """Ajusta `project_name`/`key` de um pacote já acumulado — usado no modo
        "relatório único", onde o pacote é criado sob o sentinel
        SINGLE_PACKAGE_KEY (nome real só fica conhecido depois de percorrer
        todas as linhas) e precisa ser renomeado ao final. Não faz nada se
        `package_key` nunca foi acumulado (ex: nenhuma linha válida)."""
        pkg = self._packages.get(package_key)
        if pkg is not None:
            pkg.project_name = name
            pkg.key = name

    def build(self) -> list[WorkPackage]:
        result = []
        for key in self._package_order:
            pkg = self._packages[key]
            pkg.groups = [
                self._groups_by_package[key][name] for name in self._group_order_by_package[key]
            ]
            result.append(pkg)
        return result


def _classify_incomplete_row(
    row_number: int, hs_value, obs_filled: bool, hs_filled: bool, dados_filled: bool
) -> RowIssue | None:
    """Decide se uma linha com Hs/Observação parcialmente preenchidos é um
    apontamento incompleto de verdade (devolve um RowIssue) ou deve ser
    ignorada silenciosamente — só é chamada quando pelo menos um dos dois
    campos está vazio.

    Não reporta como incompleta: uma linha totalmente em branco (comum no
    fim da planilha exportada), boilerplate de rodapé/assinatura (ex: uma
    fileira de "_______" ou os rótulos "Supervisor"/"Colaborador"/"Gerente",
    que às vezes caem na coluna Hs por causa do layout do export), ou uma
    linha de subtotal que o Projectile insere sozinho (só "Hs" preenchido e
    nem a Data presente, ex: "98,63" sozinho) — nenhuma dessas é um
    apontamento real esquecido pela metade.
    """
    is_real_partial = False
    if obs_filled and not hs_filled:
        is_real_partial = True
    elif hs_filled and not obs_filled:
        try:
            _parse_hs_value(hs_value)
            is_real_partial = dados_filled
        except ValueError:
            is_real_partial = False

    if not is_real_partial:
        return None
    falta = "Observação" if not obs_filled else "Hs"
    return RowIssue(
        row=row_number,
        reason="dados_incompletos",
        message=f"Linha {row_number}: apontamento incompleto, falta preencher \"{falta}\".",
    )


def _find_header_row(ws) -> int:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        values = [c.value for c in row]
        if values[:3] == ["Dados", "Horário", "Hs"]:
            return row[0].row
    raise ValueError("Não encontrei a linha de cabeçalho 'Dados | Horário | Hs' na planilha.")


def parse_projectile_export(
    file_path: str, split_by_package: bool = False
) -> tuple[list[WorkPackage], list[RowIssue]]:
    """Lê o export do Projectile e agrupa as horas por Prefixo/Descrição da coluna Observação.

    Se `split_by_package` for True, separa as linhas por pacote de trabalho (projeto)
    primeiro, gerando um WorkPackage por valor distinto da coluna "Pacote de Trabalho"
    — usado no modo "múltiplos relatórios". Se for False (padrão, modo "relatório
    único"), ignora essa coluna e tudo cai num único pacote, cujo nome vem da coluna
    "Projeto" (comportamento histórico do app).

    Também retorna a lista de linhas ignoradas por parecerem um apontamento incompleto
    ou mal formatado (Hs/Observação parcialmente preenchidos, falta o "-" ou "_" separando
    prefixo/descrição, descrição vazia, ou Hs não numérico) — para exibir um aviso na
    tela. Linhas totalmente em branco (Hs e Observação ambos vazios) não entram nessa
    lista por serem comuns no fim da planilha exportada, não um erro do usuário — assim
    como linhas de rodapé/assinatura (ex: uma fileira de "_______" ou os rótulos
    "Supervisor"/"Colaborador"/"Gerente"), que também não são apontamento incompleto.

    Levanta ValueError se a planilha não tiver as colunas obrigatórias ("Hs",
    "Observação") mesmo depois de encontrar a linha de cabeçalho — em vez de deixar
    vazar um KeyError bruto.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.worksheets[0]

    header_row = _find_header_row(ws)
    col_index: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        name = ws.cell(row=header_row, column=c).value
        # em caso de coluna duplicada no cabeçalho, mantém a primeira ocorrência
        if name not in col_index:
            col_index[name] = c

    missing_cols = [col for col in ("Hs", "Observação") if col not in col_index]
    if missing_cols:
        raise ValueError(
            "Não encontrei a(s) coluna(s) "
            + ", ".join(f'"{col}"' for col in missing_cols)
            + " na planilha. Confira se é o export correto do Projectile."
        )

    hs_col = col_index["Hs"]
    obs_col = col_index["Observação"]
    dados_col = col_index.get("Dados")
    projeto_col = col_index.get("Projeto")
    pacote_col = col_index.get("Pacote de Trabalho") if split_by_package else None

    accumulator = _PackageAccumulator()
    single_project_name = ""
    issues: list[RowIssue] = []

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        row_number = row[0].row

        if not split_by_package and not single_project_name and projeto_col:
            projeto_value = row[projeto_col - 1].value
            if projeto_value:
                single_project_name = str(projeto_value).strip()

        hs_value = row[hs_col - 1].value
        obs_value = row[obs_col - 1].value
        obs_stripped = str(obs_value).strip() if obs_value is not None else ""
        obs_filled = bool(obs_stripped)
        hs_filled = hs_value not in (None, "")

        if not obs_filled or not hs_filled:
            dados_filled = bool(row[dados_col - 1].value) if dados_col else True
            issue = _classify_incomplete_row(row_number, hs_value, obs_filled, hs_filled, dados_filled)
            if issue:
                issues.append(issue)
            continue

        obs_value = obs_stripped
        separator_match = re.search(r"[-_]", obs_value)
        if not separator_match:
            issues.append(RowIssue(
                row=row_number,
                reason="sem_separador",
                message=f"Linha {row_number}: Observação \"{obs_value}\" sem \"-\" ou \"_\" separando prefixo e descrição.",
            ))
            continue
        sep_index = separator_match.start()
        prefix, description = obs_value[:sep_index], obs_value[sep_index + 1:]
        prefix = prefix.strip()
        description = description.strip()
        if not prefix or not description:
            if not prefix:
                vazio = "prefixo vazio"
            else:
                vazio = "descrição vazia"
            issues.append(RowIssue(
                row=row_number,
                reason="descricao_vazia",
                message=f"Linha {row_number}: {vazio} em \"{obs_value}\".",
            ))
            continue

        try:
            hs_float = round(_parse_hs_value(hs_value), 3)
        except ValueError:
            issues.append(RowIssue(
                row=row_number,
                reason="hs_invalido",
                message=f"Linha {row_number}: valor de Hs \"{hs_value}\" não é um número válido.",
            ))
            continue

        if split_by_package:
            pacote_value = row[pacote_col - 1].value if pacote_col else None
            package_key = _extract_package_key(pacote_value)
            if not package_key:
                # Não dá pra identificar o projeto dessa linha — a chave de fallback
                # sempre inclui o número da linha, então NUNCA é compartilhada entre
                # linhas diferentes (mesmo que o texto bruto de "Pacote de Trabalho"
                # coincida, ex: duas linhas com "N/A"). Confiar em texto bruto igual
                # como sinal de "mesmo projeto" já causou a mesclagem silenciosa de
                # projetos diferentes uma vez — mais seguro separar sempre e deixar o
                # usuário juntar manualmente se, de fato, forem o mesmo projeto.
                raw_pacote = str(pacote_value).strip() if pacote_value else ""
                package_key = f"{raw_pacote} (linha {row_number})" if raw_pacote else f"Geral (linha {row_number})"
                issues.append(RowIssue(
                    row=row_number,
                    reason="pacote_nao_identificado",
                    message=(
                        f"Linha {row_number}: não consegui identificar o projeto a partir de "
                        f"\"Pacote de Trabalho\" (\"{raw_pacote}\") — agrupado separadamente."
                    ),
                ))
        else:
            package_key = SINGLE_PACKAGE_KEY

        accumulator.add_activity(
            package_key=package_key,
            package_name=package_key,
            group_name=prefix,
            description=description,
            hours=hs_float,
        )

    if not split_by_package:
        # Se a coluna "Projeto" não existir ou vier vazia em todas as linhas, deixa
        # o nome do pacote realmente vazio em vez de vazar o sentinel interno
        # SINGLE_PACKAGE_KEY para a tela/relatório final — um campo vazio é óbvio
        # de notar e corrigir, um placeholder de implementação não é.
        accumulator.rename_package(SINGLE_PACKAGE_KEY, single_project_name)

    return accumulator.build(), issues
