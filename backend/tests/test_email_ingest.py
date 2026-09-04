"""Trava o comportamento atual de `backend/app/email_ingest.py` — o módulo
que alimenta o KPI executivo de "Horas Faturadas" sem revisão humana (ver
docstring do módulo). Qualquer regressão aqui é, por definição, um número
executivo errado indo pro painel de gerência silenciosamente.

As fixtures .xlsx usadas contra os caminhos felizes são geradas por
`generator.generate_report` (via `tests/helpers.make_report`), o mesmo
caminho de código de produção que os relatórios reais do app percorrem —
não um mock da estrutura do arquivo. Os casos negativos (fórmula não
reconhecida, label ausente) usam `openpyxl.Workbook()` manual, porque
precisam representar exatamente uma forma que o `generator` NUNCA produz.
"""
from __future__ import annotations

from datetime import date, datetime, timezone

import pytest
from openpyxl import Workbook

from backend.app.email_ingest import (
    EmailIngestError,
    _dedupe_by_stem,
    compute_business_days_elapsed,
    match_project,
    read_pacote_scope,
    read_pdf_report_data,
    read_project_identity,
    resolve_total_hours,
)
from backend.app.generator import ActivityInput, GroupInput

from .helpers import make_report, make_report_pdf


# ---------------------------------------------------------------------------
# resolve_total_hours — caminho feliz
# ---------------------------------------------------------------------------

def test_resolve_total_hours_single_group_single_activity(tmp_path):
    groups = [GroupInput(name="Grupo Único", performance=1.0, activities=[ActivityInput("Única Ativ", 7.5)])]
    path = make_report(tmp_path, month_label="Setembro/2026", groups=groups)

    hours, month_label = resolve_total_hours(path)

    assert hours == pytest.approx(7.5, abs=1e-3)
    assert month_label == "Setembro/2026"


def test_resolve_total_hours_multiple_groups_different_performance(tmp_path):
    groups = [
        GroupInput(
            name="Grupo A",
            performance=1.1,
            activities=[ActivityInput("Ativ 1", 10.0), ActivityInput("Ativ 2", 5.0)],
        ),
        GroupInput(
            name="Grupo B",
            performance=0.9,
            activities=[ActivityInput("Ativ 3", 8.0)],
        ),
    ]
    path = make_report(tmp_path, month_label="Julho/2026", groups=groups)

    hours, month_label = resolve_total_hours(path)

    expected = round(15.0 * 1.1 + 8.0 * 0.9, 3)  # bruto_total * performance, somado entre grupos
    assert hours == pytest.approx(expected, abs=1e-3)
    assert month_label == "Julho/2026"


def test_resolve_total_hours_group_without_real_activities_resolves_to_zero(tmp_path):
    """Grupo sem nenhuma atividade real dispara o fallback "(sem atividades
    apontadas)" em `generator._build_groups_xml` — esse grupo não entra na
    soma total (nenhuma célula C dele é adicionada ao SUM), então deve
    contribuir 0 para o total resolvido."""
    groups = [
        GroupInput(name="Grupo Vazio", performance=1.0, activities=[]),
        GroupInput(name="Grupo Com Horas", performance=1.2, activities=[ActivityInput("Ativ X", 4.0)]),
    ]
    path = make_report(tmp_path, month_label="Agosto/2026", groups=groups)

    hours, month_label = resolve_total_hours(path)

    expected = round(4.0 * 1.2, 3)  # Grupo Vazio contribui 0
    assert hours == pytest.approx(expected, abs=1e-3)
    assert month_label == "Agosto/2026"


def test_resolve_total_hours_mixed_real_and_extra_activities(tmp_path):
    """Atividades extras (hours=None) misturadas com reais não entram no
    bruto_total nem na soma final — só as reais contam."""
    groups = [
        GroupInput(
            name="Grupo Misto",
            performance=1.0,
            activities=[
                ActivityInput("Ativ Real 1", 6.0),
                ActivityInput("Extra sem apontamento", None),
                ActivityInput("Ativ Real 2", 2.0),
            ],
        )
    ]
    path = make_report(tmp_path, month_label="Outubro/2026", groups=groups)

    hours, month_label = resolve_total_hours(path)

    expected = round(8.0 * 1.0, 3)  # bruto_total = 6.0 + 2.0 (extras não contam)
    assert hours == pytest.approx(expected, abs=1e-3)


# ---------------------------------------------------------------------------
# resolve_total_hours — caminhos negativos
# ---------------------------------------------------------------------------

def test_resolve_total_hours_raises_when_label_missing(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws["B1"] = "Nada relevante aqui"
    path = str(tmp_path / "sem_label.xlsx")
    wb.save(path)

    with pytest.raises(EmailIngestError):
        resolve_total_hours(path)


@pytest.mark.parametrize(
    "formula",
    [
        "=E5+F5",  # soma direta entre células, não SUM(...)
        "=MAX(E5,F5)",  # função não reconhecida
        "=E5*F5*G5",  # produto de 3 células, não das 2 formas reconhecidas
    ],
)
def test_resolve_total_hours_raises_on_unrecognized_formula_shape(tmp_path, formula):
    wb = Workbook()
    ws = wb.active
    ws["B5"] = "Total de horas Julho/2026:"
    ws["C5"] = formula
    path = str(tmp_path / "formula_desconhecida.xlsx")
    wb.save(path)

    with pytest.raises(EmailIngestError):
        resolve_total_hours(path)


# ---------------------------------------------------------------------------
# read_project_identity
# ---------------------------------------------------------------------------

def test_read_project_identity_matches_header_written_by_generator(tmp_path):
    groups = [GroupInput(name="Grupo", performance=1.0, activities=[ActivityInput("Ativ", 1.0)])]
    path = make_report(
        tmp_path,
        project_code="1546.6.4",
        project_name="Sangam - Cabina Bruta",
        groups=groups,
    )

    project_code, project_name = read_project_identity(path)

    assert project_code == "1546.6.4"
    assert project_name == "Sangam - Cabina Bruta"


# ---------------------------------------------------------------------------
# read_pacote_scope
# ---------------------------------------------------------------------------

def test_read_pacote_scope_none_by_default(tmp_path):
    """Sem `pacote_scope` na geração, o relatório cobre o projeto inteiro —
    `read_pacote_scope` devolve None."""
    groups = [GroupInput(name="Grupo", performance=1.0, activities=[ActivityInput("Ativ", 1.0)])]
    path = make_report(tmp_path, groups=groups)

    assert read_pacote_scope(path) is None


def test_read_pacote_scope_matches_marker_written_by_generator(tmp_path):
    groups = [GroupInput(name="Grupo", performance=1.0, activities=[ActivityInput("Ativ", 1.0)])]
    pacote_text = "1546.6.4-002 Legislation Package - Sangam - Cabina Bruta - 07.2026"
    path = make_report(tmp_path, groups=groups, pacote_scope=pacote_text)

    assert read_pacote_scope(path) == pacote_text


# ---------------------------------------------------------------------------
# read_pdf_report_data — equivalente pro .pdf de resolve_total_hours +
# read_project_identity + read_pacote_scope juntos (sem fórmula/célula, o
# .pdf carrega tudo isso como metadado, ver pdf_generator._embed_report_metadata)
# ---------------------------------------------------------------------------

def test_read_pdf_report_data_matches_metadata_written_by_generator(tmp_path):
    groups = [
        GroupInput(name="Grupo A", performance=1.1, activities=[ActivityInput("Ativ 1", 10.0)]),
        GroupInput(name="Grupo B", performance=0.9, activities=[ActivityInput("Ativ 2", 8.0)]),
    ]
    pdf_path = make_report_pdf(
        tmp_path,
        project_code="1546.6.4",
        project_name="Sangam - Cabina Bruta",
        month_label="Julho/2026",
        groups=groups,
    )

    data = read_pdf_report_data(pdf_path)

    assert data["project_code"] == "1546.6.4"
    assert data["project_name"] == "Sangam - Cabina Bruta"
    assert data["month_label"] == "Julho/2026"
    assert data["total_hours"] == pytest.approx(10.0 * 1.1 + 8.0 * 0.9, abs=1e-3)
    assert data["pacote_scope"] is None


def test_read_pdf_report_data_matches_pacote_scope_when_set(tmp_path):
    groups = [GroupInput(name="Grupo", performance=1.0, activities=[ActivityInput("Ativ", 1.0)])]
    pacote_text = "1546.6.4-002 Legislation Package - Sangam - Cabina Bruta - 07.2026"
    pdf_path = make_report_pdf(tmp_path, groups=groups, pacote_scope=pacote_text)

    assert read_pdf_report_data(pdf_path)["pacote_scope"] == pacote_text


def test_read_pdf_report_data_raises_when_pdf_was_not_generated_by_this_app(tmp_path):
    """PDF real, mas sem o metadado que só `pdf_generator.py` grava — ex: um
    PDF qualquer que o cliente/gerente anexou por engano."""
    from reportlab.pdfgen import canvas

    path = str(tmp_path / "outro.pdf")
    c = canvas.Canvas(path)
    c.drawString(100, 750, "não é um relatório desta automação")
    c.save()

    with pytest.raises(EmailIngestError):
        read_pdf_report_data(path)


# ---------------------------------------------------------------------------
# _dedupe_by_stem — mesmo relatório em .xlsx e .pdf na mesma mensagem só deve
# virar 1 amostra (billed_hours é SOMADO entre amostras em
# management.compute_monthly_kpis — processar os dois dobraria a hora)
# ---------------------------------------------------------------------------

def test_dedupe_by_stem_prefers_xlsx_when_pdf_listed_first(tmp_path):
    xlsx_file = tmp_path / "a.xlsx"
    pdf_file = tmp_path / "a.pdf"
    xlsx_file.write_bytes(b"xlsx")
    pdf_file.write_bytes(b"pdf")

    survivors = _dedupe_by_stem([(str(pdf_file), "Relatorio.pdf"), (str(xlsx_file), "Relatorio.xlsx")])

    assert survivors == [str(xlsx_file)]
    assert xlsx_file.exists()
    assert not pdf_file.exists()  # o perdedor do dedup é removido do disco


def test_dedupe_by_stem_prefers_xlsx_when_xlsx_listed_first(tmp_path):
    xlsx_file = tmp_path / "a.xlsx"
    pdf_file = tmp_path / "a.pdf"
    xlsx_file.write_bytes(b"xlsx")
    pdf_file.write_bytes(b"pdf")

    survivors = _dedupe_by_stem([(str(xlsx_file), "Relatorio.xlsx"), (str(pdf_file), "Relatorio.pdf")])

    assert survivors == [str(xlsx_file)]
    assert not pdf_file.exists()


def test_dedupe_by_stem_keeps_unrelated_stems_separate(tmp_path):
    file_a = tmp_path / "a.xlsx"
    file_b = tmp_path / "b.pdf"
    file_a.write_bytes(b"a")
    file_b.write_bytes(b"b")

    survivors = _dedupe_by_stem([(str(file_a), "Projeto A.xlsx"), (str(file_b), "Projeto B.pdf")])

    assert sorted(survivors) == sorted([str(file_a), str(file_b)])
    assert file_a.exists() and file_b.exists()


# ---------------------------------------------------------------------------
# match_project
# ---------------------------------------------------------------------------

def test_match_project_exact_match_scores_one(tmp_path):
    candidates = {"1": "Sangam - Cabina Bruta", "2": "Projeto Alpha", "3": "Projeto Beta"}

    project_id, project_name, score = match_project("Sangam - Cabina Bruta", candidates)

    assert project_id == "1"
    assert project_name == "Sangam - Cabina Bruta"
    assert score == 1.0


def test_match_project_similar_but_not_exact_returns_closest_candidate(tmp_path):
    candidates = {"1": "Sangam - Cabina Bruta", "2": "Projeto Alpha", "3": "Projeto Beta"}

    project_id, project_name, score = match_project("Projeto Alfa", candidates)

    assert project_id == "2"
    assert project_name == "Projeto Alpha"
    assert 0.5 < score < 1.0


def test_match_project_raises_when_no_candidates():
    with pytest.raises(EmailIngestError):
        match_project("Qualquer coisa", {})


# ---------------------------------------------------------------------------
# compute_business_days_elapsed
# ---------------------------------------------------------------------------

def test_compute_business_days_elapsed_known_case():
    # Caso já validado manualmente nesta sessão:
    # count_business_days(date(2026,7,31), date(2026,8,5)) == 3.
    # compute_business_days_elapsed ancora no ÚLTIMO dia do mês do relatório
    # (31/07/2026), então o mesmo intervalo reaparece aqui usando 05/08/2026
    # como data de envio do e-mail.
    sent_at = datetime(2026, 8, 5, 12, 0, tzinfo=timezone.utc)

    result = compute_business_days_elapsed("Julho/2026", sent_at)

    assert result == 3


def test_compute_business_days_elapsed_crosses_fixed_national_holiday():
    """Novembro/2026 fecha em 30/11 (segunda) — usado como âncora. 25/12/2026
    é uma sexta-feira (dia útil "normal" se não fosse feriado nacional fixo,
    ver `generator._national_holidays`). Prova de que o feriado é excluído:
    contar até 24/12 e até 25/12 dá o MESMO resultado — se 25/12 fosse
    contado como dia útil comum, o segundo valor seria 1 a mais."""
    result_before_holiday = compute_business_days_elapsed(
        "Novembro/2026", datetime(2026, 12, 24, 9, 0, tzinfo=timezone.utc)
    )
    result_including_holiday = compute_business_days_elapsed(
        "Novembro/2026", datetime(2026, 12, 25, 9, 0, tzinfo=timezone.utc)
    )

    assert result_before_holiday == 18
    assert result_including_holiday == 18  # 25/12 (sexta) não soma por ser feriado


def test_compute_business_days_elapsed_weekend_only_span():
    """Julho/2026 fecha em 31/07 (sexta-feira) — usado como âncora. Enviado
    em 03/08/2026 (segunda seguinte), cruzando só um fim de semana (01/08
    sábado, 02/08 domingo) sem nenhum feriado no meio -> só a segunda-feira
    conta como dia útil."""
    sent_at = datetime(2026, 8, 3, 10, 0, tzinfo=timezone.utc)

    result = compute_business_days_elapsed("Julho/2026", sent_at)

    assert result == 1


def test_compute_business_days_elapsed_invalid_label_raises():
    with pytest.raises(EmailIngestError):
        compute_business_days_elapsed("não é uma competência válida", datetime.now(timezone.utc))
