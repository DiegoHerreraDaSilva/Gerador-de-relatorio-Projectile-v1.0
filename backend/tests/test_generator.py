"""Trava a FORMA do XML gerado por `generate_report` — especificamente as 3
formas de fórmula que `email_ingest._resolve_cell` reconhece (SUM(...),
referência simples, produto A*B). `email_ingest.resolve_total_hours` depende
disso continuar exatamente assim; se uma refatoração futura de
`_build_groups_xml` mudar uma dessas formas sem querer, este teste E os de
`test_email_ingest.py` devem falhar juntos — essa é a rede de segurança
pedida para a próxima fase."""
from __future__ import annotations

import re

import pytest
from openpyxl import load_workbook

from backend.app.generator import ActivityInput, GroupInput, HIDDEN_HELPER_COL

from .helpers import make_report

SIMPLE_REF_RE = re.compile(r"^[A-Z]+\d+$")
SUM_RE = re.compile(r"^SUM\(.+\)$")


@pytest.fixture
def report_path(tmp_path):
    groups = [
        GroupInput(
            name="Grupo A",
            performance=1.1,
            activities=[
                ActivityInput("Ativ 1", 10.0),
                ActivityInput("Extra sem apontamento", None),
            ],
        ),
        GroupInput(
            name="Grupo B",
            performance=0.9,
            activities=[ActivityInput("Ativ 2", 8.0)],
        ),
    ]
    return make_report(
        tmp_path,
        project_code="1546.6.4",
        project_name="Sangam - Cabina Bruta",
        month_label="Julho/2026",
        groups=groups,
    )


def _find_total_label_cell(ws):
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("Total de horas"):
                return cell
    raise AssertionError("Célula 'Total de horas ...:' não encontrada no XML gerado.")


def test_total_label_is_in_column_b(report_path):
    wb = load_workbook(report_path, data_only=False)
    ws = wb.active

    label_cell = _find_total_label_cell(ws)

    assert label_cell.column_letter == "B"
    assert re.match(r"^Total de horas .+:$", label_cell.value)


def test_total_formula_is_sum_of_refs(report_path):
    wb = load_workbook(report_path, data_only=False)
    ws = wb.active

    label_cell = _find_total_label_cell(ws)
    total_formula_cell = ws[f"C{label_cell.row}"]

    assert isinstance(total_formula_cell.value, str) and total_formula_cell.value.startswith("=")
    formula = total_formula_cell.value[1:]
    assert SUM_RE.match(formula), f"Fórmula do total não bate com ^SUM(.+)$: {formula!r}"


def test_group_subtotal_cells_are_simple_references_to_literal_products(report_path):
    """Cada célula referenciada dentro do SUM(...) do total deve ser, ela
    mesma, uma fórmula de referência simples (ex: "=E17"), que por sua vez
    aponta pra um NÚMERO LITERAL já calculado (bruto * performance) — não uma
    fórmula "=E{linha}*F{linha}": a coluna F (Performance) fica em branco no
    relatório final (não pode aparecer pro destinatário), então o produto é
    calculado em Python e escrito como valor, sem depender de F."""
    wb = load_workbook(report_path, data_only=False)
    ws = wb.active

    label_cell = _find_total_label_cell(ws)
    total_formula_cell = ws[f"C{label_cell.row}"]
    sum_inner = total_formula_cell.value[1:][len("SUM(") : -1]
    refs = [r.strip() for r in sum_inner.split(",") if r.strip()]

    assert len(refs) == 2  # Grupo A e Grupo B contribuem, cada um com 1 ref

    expected_products = {10.0 * 1.1, 8.0 * 0.9}  # bruto * performance de cada grupo do fixture
    for ref in refs:
        subtotal_cell = ws[ref]
        assert isinstance(subtotal_cell.value, str) and subtotal_cell.value.startswith("=")
        subtotal_formula = subtotal_cell.value[1:]
        assert SIMPLE_REF_RE.match(subtotal_formula), (
            f"Fórmula de subtotal do grupo não é referência simples (^[A-Z]+\\d+$): {subtotal_formula!r}"
        )

        product_cell = ws[subtotal_formula]
        assert isinstance(product_cell.value, (int, float)), (
            f"Célula referenciada deveria ser um número literal, veio {product_cell.value!r}"
        )
        assert any(abs(product_cell.value - expected) < 1e-6 for expected in expected_products)

        # coluna F (Performance) da linha de atividade (mesma linha do
        # subtotal em C) deve estar vazia — é o que "não pode aparecer a
        # performance" exige.
        perf_col_cell = ws[f"F{subtotal_cell.row}"]
        assert perf_col_cell.value is None


def test_header_cells_contain_exact_project_code_and_name(report_path):
    wb = load_workbook(report_path, data_only=False)
    ws = wb.active

    assert ws["B8"].value == "1546.6.4"
    assert ws["C9"].value == "Sangam - Cabina Bruta"


def test_pacote_scope_marker_is_empty_by_default(tmp_path):
    """Sem `pacote_scope`, o relatório cobre o projeto inteiro — a marca
    oculta (logo abaixo do "Total de horas ...:", ver `_build_totals_row`)
    deve ficar vazia."""
    groups = [GroupInput(name="Grupo A", performance=1.0, activities=[ActivityInput("Ativ 1", 10.0)])]
    path = make_report(tmp_path, groups=groups)
    wb = load_workbook(path, data_only=False)
    ws = wb.active

    label_cell = _find_total_label_cell(ws)
    marker_cell = ws[f"{HIDDEN_HELPER_COL}{label_cell.row + 1}"]
    assert marker_cell.value in (None, "")


def test_pacote_scope_marker_carries_pacote_text(tmp_path):
    """Com `pacote_scope`, a marca oculta guarda o texto do pacote coberto —
    é isso que `email_ingest.read_pacote_scope` lê de volta pra impedir que
    um relatório de 1 pacote marque o projeto inteiro como enviado."""
    groups = [GroupInput(name="Grupo A", performance=1.0, activities=[ActivityInput("Ativ 1", 10.0)])]
    pacote_text = "1546.6.4-002 Legislation Package - Sangam - Cabina Bruta - 07.2026"
    path = make_report(tmp_path, groups=groups, pacote_scope=pacote_text)
    wb = load_workbook(path, data_only=False)
    ws = wb.active

    label_cell = _find_total_label_cell(ws)
    marker_cell = ws[f"{HIDDEN_HELPER_COL}{label_cell.row + 1}"]
    assert marker_cell.value == pacote_text
