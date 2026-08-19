"""Leitura e agrupamento do relatório exportado do Projectile."""
from __future__ import annotations

import re
from dataclasses import dataclass, field

import openpyxl


@dataclass
class Activity:
    description: str
    hours: float


@dataclass
class Group:
    name: str
    activities: list[Activity] = field(default_factory=list)

    @property
    def total_hours(self) -> float:
        return round(sum(a.hours for a in self.activities), 3)


@dataclass
class WorkPackage:
    key: str
    project_name: str
    groups: list[Group] = field(default_factory=list)


@dataclass
class RowIssue:
    row: int
    reason: str
    message: str


# variantes de travessão/hífen que aparecem em exports reais no lugar do hífen ASCII
# comum ("-") — normalizadas antes de separar os segmentos de "Pacote de Trabalho".
_DASH_VARIANTS = ["–", "—", "−", "－"]  # en dash, em dash, minus sign, fullwidth hyphen


def _extract_package_key(raw_value) -> str | None:
    """Extrai o identificador do projeto a partir da coluna "Pacote de Trabalho".

    O valor vem no formato "<código>-<seq> Legislation Package - <Projeto> - <resto>"
    (ex: "1546.6.4-002 Legislation Package - Sangam - Cabina Bruta - ..."). O separador
    dos segmentos é " - " (espaço-hífen-espaço), não hífen puro — usar hífen puro
    quebraria nomes de projeto que já contêm hífen, como "Para-barro". O nome do
    projeto é o 2º segmento — é isso que agrupa várias linhas de "Pacote de Trabalho"
    diferentes (6.4-001, 6.4-002, ...) num único relatório ("Sangam").

    Retorna None quando o valor NÃO segue essa estrutura (menos de 2 segmentos depois
    de separar por " - ") — ex: "N/A", "TBD", "-", ou qualquer texto livre. Antes essa
    situação caía num fallback ingênuo que devolvia o texto bruto como se fosse um nome
    de projeto confiável; o problema é que duas linhas de projetos genuinamente
    diferentes podem compartilhar o mesmo texto de preenchimento genérico (ambas com
    "N/A", por exemplo) e acabariam sendo mescladas num único pacote sem nenhum aviso.
    Devolver None força o chamador a tratar como "não identificado" (chave própria por
    linha + aviso) em vez de confiar cegamente no texto.
    """
    if not raw_value:
        return None
    # separador tolera espaçamento irregular (às vezes vem "- " ou " -" só de um lado,
    # ou espaço duplo), normaliza non-breaking space (\xa0) e variantes de travessão
    # unicode (en/em dash, sinal de menos, hífen fullwidth), comuns em exports do Excel.
    normalized = str(raw_value).replace("\xa0", " ")
    for dash in _DASH_VARIANTS:
        normalized = normalized.replace(dash, "-")
    # o padrão exige espaço em pelo menos um dos lados do hífen (não precisa ser dos
    # dois) — só assim ele conta como separador de segmento; um hífen sem espaço
    # nenhum ao redor (ex: "Para-barro") nunca é separador. Um padrão assimétrico
    # (só exigindo espaço de um lado fixo) deixaria passar despercebido um separador
    # real que só tem espaço do outro lado, quebrando o texto no lugar errado.
    parts = re.split(r"\s+-\s*|\s*-\s+", normalized)
    if len(parts) >= 2:
        return parts[1].strip() or None
    return None


def _parse_hs_value(value) -> float:
    """Converte o valor de Hs para float, aceitando tanto ponto quanto vírgula como
    separador decimal — exports variam conforme a configuração regional de quem gerou
    a planilha, e um valor como "2,5" é um apontamento real, não deveria ser tratado
    como texto de rodapé só porque float() sozinho não entende vírgula.
    """
    try:
        return float(value)
    except (TypeError, ValueError):
        pass
    if isinstance(value, str):
        try:
            return float(value.strip().replace(",", "."))
        except ValueError:
            pass
    raise ValueError(f"valor de Hs não numérico: {value!r}")


def _find_header_row(ws) -> int:
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        values = [c.value for c in row]
        if values[:3] == ["Dados", "Horário", "Hs"]:
            return row[0].row
    raise ValueError("Não encontrei a linha de cabeçalho 'Dados | Horário | Hs' na planilha.")


def parse_projectile_export(
    file_path: str, split_by_package: bool = False
) -> tuple[list[WorkPackage], list[RowIssue]]:
    """Lê o export do Projectile e agrupa as horas por Prefixo/Descrição da coluna Observação.

    Se `split_by_package` for True, separa as linhas por pacote de trabalho (projeto)
    primeiro, gerando um WorkPackage por valor distinto da coluna "Pacote de Trabalho"
    — usado no modo "múltiplos relatórios". Se for False (padrão, modo "relatório
    único"), ignora essa coluna e tudo cai num único pacote, cujo nome vem da coluna
    "Projeto" (comportamento histórico do app).

    Também retorna a lista de linhas ignoradas por parecerem um apontamento incompleto
    ou mal formatado (Hs/Observação parcialmente preenchidos, falta o "_" separando
    prefixo/descrição, descrição vazia, ou Hs não numérico) — para exibir um aviso na
    tela. Linhas totalmente em branco (Hs e Observação ambos vazios) não entram nessa
    lista por serem comuns no fim da planilha exportada, não um erro do usuário — assim
    como linhas de rodapé/assinatura (ex: uma fileira de "_______" ou os rótulos
    "Supervisor"/"Colaborador"/"Gerente"), que também não são apontamento incompleto.

    Levanta ValueError se a planilha não tiver as colunas obrigatórias ("Hs",
    "Observação") mesmo depois de encontrar a linha de cabeçalho — em vez de deixar
    vazar um KeyError bruto.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.worksheets[0]

    header_row = _find_header_row(ws)
    col_index: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        name = ws.cell(row=header_row, column=c).value
        # em caso de coluna duplicada no cabeçalho, mantém a primeira ocorrência
        if name not in col_index:
            col_index[name] = c

    missing_cols = [col for col in ("Hs", "Observação") if col not in col_index]
    if missing_cols:
        raise ValueError(
            "Não encontrei a(s) coluna(s) "
            + ", ".join(f'"{col}"' for col in missing_cols)
            + " na planilha. Confira se é o export correto do Projectile."
        )

    hs_col = col_index["Hs"]
    obs_col = col_index["Observação"]
    dados_col = col_index.get("Dados")
    projeto_col = col_index.get("Projeto")
    pacote_col = col_index.get("Pacote de Trabalho") if split_by_package else None

    SINGLE_PACKAGE_KEY = "__single__"
    packages: dict[str, WorkPackage] = {}
    package_order: list[str] = []
    groups_by_package: dict[str, dict[str, Group]] = {}
    group_order_by_package: dict[str, list[str]] = {}
    single_project_name = ""
    issues: list[RowIssue] = []

    for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
        row_number = row[0].row

        if not split_by_package and not single_project_name and projeto_col:
            projeto_value = row[projeto_col - 1].value
            if projeto_value:
                single_project_name = str(projeto_value).strip()

        hs_value = row[hs_col - 1].value
        obs_value = row[obs_col - 1].value
        obs_stripped = str(obs_value).strip() if obs_value is not None else ""
        obs_filled = bool(obs_stripped)
        hs_filled = hs_value not in (None, "")

        if not obs_filled or not hs_filled:
            # só é reportado como "apontamento incompleto" quando o lado preenchido
            # parece dado real de timesheet — não boilerplate de rodapé/assinatura
            # (ex: uma linha de "_______" ou os rótulos "Supervisor"/"Colaborador",
            # que às vezes caem na coluna Hs por causa do layout do export).
            is_real_partial = False
            if obs_filled and not hs_filled:
                is_real_partial = True
            elif hs_filled and not obs_filled:
                dados_filled = bool(row[dados_col - 1].value) if dados_col else True
                try:
                    _parse_hs_value(hs_value)
                    # Uma linha com só "Hs" preenchido e nem a Data presente
                    # provavelmente é uma linha de subtotal/soma que o Projectile
                    # insere na planilha (ex: "98,63" sozinho, sem nenhum outro
                    # dado) — não um apontamento real esquecido pela metade.
                    is_real_partial = dados_filled
                except ValueError:
                    is_real_partial = False
            if is_real_partial:
                falta = "Observação" if not obs_filled else "Hs"
                issues.append(RowIssue(
                    row=row_number,
                    reason="dados_incompletos",
                    message=f"Linha {row_number}: apontamento incompleto, falta preencher \"{falta}\".",
                ))
            continue

        obs_value = obs_stripped
        if "_" not in obs_value:
            issues.append(RowIssue(
                row=row_number,
                reason="sem_underscore",
                message=f"Linha {row_number}: Observação \"{obs_value}\" sem \"_\" separando prefixo e descrição.",
            ))
            continue
        prefix, description = obs_value.split("_", 1)
        prefix = prefix.strip()
        description = description.strip()
        if not prefix or not description:
            if not prefix:
                vazio = "prefixo vazio"
            else:
                vazio = "descrição vazia"
            issues.append(RowIssue(
                row=row_number,
                reason="descricao_vazia",
                message=f"Linha {row_number}: {vazio} em \"{obs_value}\".",
            ))
            continue

        try:
            hs_float = round(_parse_hs_value(hs_value), 3)
        except ValueError:
            issues.append(RowIssue(
                row=row_number,
                reason="hs_invalido",
                message=f"Linha {row_number}: valor de Hs \"{hs_value}\" não é um número válido.",
            ))
            continue

        if split_by_package:
            pacote_value = row[pacote_col - 1].value if pacote_col else None
            package_key = _extract_package_key(pacote_value)
            if not package_key:
                # Não dá pra identificar o projeto dessa linha — a chave de fallback
                # sempre inclui o número da linha, então NUNCA é compartilhada entre
                # linhas diferentes (mesmo que o texto bruto de "Pacote de Trabalho"
                # coincida, ex: duas linhas com "N/A"). Confiar em texto bruto igual
                # como sinal de "mesmo projeto" já causou a mesclagem silenciosa de
                # projetos diferentes uma vez — mais seguro separar sempre e deixar o
                # usuário juntar manualmente se, de fato, forem o mesmo projeto.
                raw_pacote = str(pacote_value).strip() if pacote_value else ""
                package_key = f"{raw_pacote} (linha {row_number})" if raw_pacote else f"Geral (linha {row_number})"
                issues.append(RowIssue(
                    row=row_number,
                    reason="pacote_nao_identificado",
                    message=(
                        f"Linha {row_number}: não consegui identificar o projeto a partir de "
                        f"\"Pacote de Trabalho\" (\"{raw_pacote}\") — agrupado separadamente."
                    ),
                ))
        else:
            package_key = SINGLE_PACKAGE_KEY

        if package_key not in packages:
            packages[package_key] = WorkPackage(key=package_key, project_name=package_key)
            package_order.append(package_key)
            groups_by_package[package_key] = {}
            group_order_by_package[package_key] = []

        groups = groups_by_package[package_key]
        group_order = group_order_by_package[package_key]
        if prefix not in groups:
            groups[prefix] = Group(name=prefix)
            group_order.append(prefix)
        group = groups[prefix]

        existing = next(
            (a for a in group.activities if a.description.casefold() == description.casefold()), None
        )
        if existing:
            existing.hours = round(existing.hours + hs_float, 3)
        else:
            group.activities.append(Activity(description=description, hours=hs_float))

    if not split_by_package and SINGLE_PACKAGE_KEY in packages:
        # Se a coluna "Projeto" não existir ou vier vazia em todas as linhas, deixa
        # o nome do pacote realmente vazio em vez de vazar o sentinel interno
        # SINGLE_PACKAGE_KEY para a tela/relatório final — um campo vazio é óbvio
        # de notar e corrigir, um placeholder de implementação não é.
        packages[SINGLE_PACKAGE_KEY].project_name = single_project_name
        packages[SINGLE_PACKAGE_KEY].key = single_project_name

    result = []
    for key in package_order:
        pkg = packages[key]
        pkg.groups = [groups_by_package[key][name] for name in group_order_by_package[key]]
        result.append(pkg)
    return result, issues
